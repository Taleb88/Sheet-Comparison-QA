import pandas as pd
import numpy as np

pd.set_option('future.no_silent_downcasting', True)

# read csv file
netflix_df = pd.read_csv('netflix_titles.csv')

# save csv file as xlsx
netflix_xlsx_df = netflix_df
netflix_xlsx_df.to_excel('netflix_titles.xlsx', index=None)

'''for x in netflix_df:
    print(x)'''

# create condensed version of file
netflix_condensed_df = pd.DataFrame()
title = netflix_df.iloc[:,2]
netflix_condensed_df['Title'] = title.copy() 
type = netflix_df.iloc[:,1]
netflix_condensed_df['Type'] = type.copy()
director = netflix_df.iloc[:,3]
netflix_condensed_df['Director'] = director.copy()
cast = netflix_df.iloc[:,4]
netflix_condensed_df['Cast'] = cast.copy()
country = netflix_df.iloc[:,5]
netflix_condensed_df['Country'] = country.copy()
date_added = netflix_df.iloc[:,6]
netflix_condensed_df['Date Added'] = date_added.copy()
release_year = netflix_df.iloc[:,7]
netflix_condensed_df['Release Year'] = release_year.copy()
rating = netflix_df.iloc[:,8]
netflix_condensed_df['Rating'] = rating.copy()
duration = netflix_df.iloc[:,9]
netflix_condensed_df['Duration'] = duration.copy()
# adding new columns to export_df
netflix_condensed_df['Media'] = ['EST'] * len(netflix_condensed_df)

# save netflix_condensed_df file
netflix_condensed_df.to_excel('netflix_condensed.xlsx', index=None)

# convert Date Added values to DateTime format
netflix_condensed_df['Date Added'] = pd.to_datetime(netflix_condensed_df['Date Added'], format='mixed')

# save netflix_condensed_df file
netflix_condensed_df.to_excel('netflix_condensed.xlsx', index=None)

# convert Date Added values to string format
netflix_condensed_df['Date Added'] = netflix_condensed_df['Date Added'].astype(str)

# save netflix_condensed_df file
netflix_condensed_df.to_excel('netflix_condensed.xlsx', index=None)

# create duplicate rows
netflix_condensed_df = pd.DataFrame(np.repeat(netflix_condensed_df.values, 2, axis=0), 
                                    columns=netflix_condensed_df.columns)

# save netflix_condensed_df file
netflix_condensed_df.to_excel('netflix_condensed.xlsx', index=None)

# replace duplicate "EST" value with "VOD" value
netflix_condensed_df['Media'] = np.where\
    (np.arange(len(netflix_condensed_df)) % 2, 'EST', 'VOD')

# save netflix_condensed_df file
netflix_condensed_df.to_excel('netflix_condensed.xlsx', index=None)

# create export_df that mirrors netflix_condensed_df
export_df = pd.DataFrame()
title = netflix_df.iloc[:,2]
export_df['Title'] = title.copy()
type = netflix_df.iloc[:,1]
export_df['Type'] = type.copy()
director = netflix_df.iloc[:,3]
export_df['Director'] = director.copy()
cast = netflix_df.iloc[:,4] 
export_df['Cast'] = cast.copy()
country = netflix_df.iloc[:,5] 
export_df['Country'] = country.copy()
date_added = netflix_df.iloc[:,6]
export_df['Date Added'] = date_added.copy()
release_year = netflix_df.iloc[:,7]
export_df['Release Year'] = release_year.copy()
rating = netflix_df.iloc[:,8]
export_df['Rating'] = rating.copy()
duration = netflix_df.iloc[:,9]
export_df['Duration'] = duration.copy()
# adding new columns to export_df
export_df['Media'] = ['EST'] * len(export_df)

# save export_df file
export_df.to_excel('export.xlsx', sheet_name='Export', index=None)

# convert Date Added values to DateTime format
export_df['Date Added'] = pd.to_datetime(export_df['Date Added'], format='mixed')

# save export_df file
export_df.to_excel('export.xlsx', index=None)

# convert Date Added values to string format
export_df['Date Added'] = export_df['Date Added'].astype(str)

# save export_df file
export_df.to_excel('export.xlsx', index=None)

# create duplicate rows
export_df = pd.DataFrame(np.repeat(export_df.values, 2, axis=0), 
                                    columns=export_df.columns)

# save export_df file
export_df.to_excel('export.xlsx', index=None)

# replace duplicate "EST" value with "VOD" value
export_df['Media'] = np.where\
    (np.arange(len(export_df)) % 2, 'EST', 'VOD')

# save export_df file
export_df.to_excel('export.xlsx', index=None)

# sort values in columns of condensed file
netflix_condensed_df = netflix_condensed_df.sort_values(by=['Type', 'Title', 'Media'])
export_df = export_df.sort_values(by=['Type', 'Title', 'Media'])

# save files
netflix_condensed_df.to_excel('netflix_condensed.xlsx', index=None)
export_df.to_excel('export.xlsx', index=None)


# ================= #
# TEST - intentionally replacing certain values to create mismatches
# ================= #
netflix_condensed_df['Release Year'] = \
    netflix_condensed_df['Release Year'].replace(1991, 1995)

# save file
netflix_condensed_df.to_excel('netflix_condensed.xlsx', index=None)

# add newly created worksheets to raw file -> netflix_titles.xlsx
with pd.ExcelWriter(
    "netflix_condensed.xlsx",
    mode="a",
    engine='openpyxl',
    if_sheet_exists="replace",
) as writer:
    export_df.to_excel(writer, sheet_name='Export', index=None)


#  
# comparing values of sheets, highlighting differences
#   compare sheet1 and export sheets in condensed file
#

import openpyxl as xl
from openpyxl.styles import PatternFill

netflix_condensed_df = xl.load_workbook("netflix_condensed.xlsx")

# highlight cells of mismatched values
fill_style = PatternFill(start_color="FE98AC", 
                         end_color="FE98AC", 
                         fill_type="solid")

# sheets used for comparison purposes
sheet1 = netflix_condensed_df['Sheet1']
export = netflix_condensed_df['Export']

for row in sheet1.iter_rows():
    for cell in row:
        current_cell_value = cell.value
        cell_location = cell.coordinate
        # if values do not match, cells get highlighted accordingly
        if current_cell_value != export[cell_location].value:
            cell.fill = fill_style

for row in export.iter_rows():
    for cell in row:
        current_cell_value = cell.value
        cell_location = cell.coordinate
        
        # if values do not match, cells get highlighted accordingly
        if current_cell_value != sheet1[cell_location].value:
            cell.fill = fill_style

netflix_condensed_df.save("netflix_condensed.xlsx")

# user can manually filter values via color when spreadsheet is open
ws = netflix_condensed_df.active
ws.auto_filter.ref = ws.dimensions

netflix_condensed_df.save("netflix_condensed.xlsx")